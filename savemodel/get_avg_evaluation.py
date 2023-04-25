import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl import load_workbook


# 定义一个函数，用于读取excel文件
def read_excel(path):
    data = pd.read_excel(path)
    return data


# 一个函数来获取源文件夹下的所有文件夹名
def get_folder_name(path):
    file_name = []
    for root, dirs, files in os.walk(path):
        for dir in dirs:
            file_name.append(os.path.join(root, dir))
    return file_name

# 一个函数来获取源文件夹下的所有文件名
def get_file_name(path):
    file_name = []
    for root, dirs, files in os.walk(path):
        for file in files:
            file_name.append(os.path.join(root, file))
    return file_name


# 一个函数读取当前文件夹下的record.xlsx文件
def read_record(path):
    # 查找当前文件夹下的record.xlsx文件，如果不存在，则报错
    if not os.path.exists(path + '/record.xlsx'):
        # 返回一个空的dataframe
        data = pd.DataFrame()
        return data
    # 查找当前文件夹下的record.xlsx文件
    else:
        # 读取record.xlsx文件
        data = pd.read_excel(path + '/record.xlsx')
        return data


# 一个函数来读取第x列往后的所有列（包括第x列）的最后n行数据，输入参数为pd的dataframe，列数，行数
def get_last_n_row(data, col, row):
    # 读取最后n行数据
    last_n_row = data.iloc[-row:, col:]
    return last_n_row


# 用一个函数来计算last_n_row每一列的平均值
def get_avg(last_n_row):
    # 读取last_n_row的列数
    col = last_n_row.shape[1]
    # 用一个空的list来存放每一列的平均值
    avg = []
    # 依次读取每一列的数据
    for i in range(col):
        # 获取第i列的数据
        col_data = last_n_row.iloc[:, i]
        # 查看他们的类型
        # print(type(col_data))
        # 将col_data转换为dataframe
        col_data = col_data.to_frame().reset_index()  # reset_index()是将原来的index转换为一列, 具体来说就是将原来的index作为一列数据，
        # 如果不想要原来的index作为一列数据，可以设置drop=True，drop是删除的意思，删除原来的index
        # 获取第i列的数据
        col_data_i = col_data.iloc[:, 1]
        col_num = []
        for idx in col_data_i.index:
            # 打印col_data_i[idx]的类型
            # print(type(col_data_i[idx]))

            col_num.append(float(col_data_i[idx]))
        # 计算col_num的平均值
        avg.append(np.mean(col_num))
    # 将avg转换为dataframe，并且不要索引
    avg_df = pd.DataFrame(avg, index=None)
    return avg_df

# 一个函数检测目标文件夹是否存在，如果不存在，则创建
def check_folder(path):
    # 检测目标文件夹是否存在
    if not os.path.exists(path):
        # 如果不存在，则创建
        os.makedirs(path)
        return True
    else:
        return False

# 一个函数来检测record有没有超过100行，如果不超过100行，则跳过
def check_record(record):
    # 获取record的行数
    row = record.shape[0]
    if row < 100:
        return False
    else:
        return True

# 用一个函数检测record.xlsx第一列中是否已经存在名为exp_name的元素，如果存在，则跳过，函数参数为record的路径名，exp_name
def check_exp_name(path, exp_name):
    # 打开record.xlsx文件
    wb = load_workbook(path + '/avg.xlsx')
    # 获取第一个sheet
    ws = wb[wb.sheetnames[0]]
    # 获取第一列的数据，ws可以理解为一个二维数
    col_data = ws['A']
    # 将col_data转换为list
    col_data = [i.value for i in col_data]
    # print(col_data)
    # 判断exp_name是否在col_data中
    if exp_name in col_data:
        # 如果存在，则返回True

        return True
    else:
        return False


# 一个函数来获取path下一级的文件夹个数
def get_folder_num(path):
    # 获取path下一级的文件夹个数
    folder_num = len(os.listdir(path))
    return folder_num



if __name__ == '__main__':
    # init 变量参数等
    # 源文件夹路径
    path = r'./'
    # 目标文件夹路径
    target_path = r'./result/'
    # 目标文件夹名称
    target_folder_name = 'AVG'

    # 一个函数检测目标文件夹是否存在，如果不存在，则创建
    if check_folder(target_path + target_folder_name):
        print('目标文件夹不存在，已创建')
    else:
        print('目标文件夹已存在')

    # 检测目标文件夹下是否存在avg.xlsx文件，如果不存在，则创建
    if not os.path.exists(target_path + target_folder_name + '/avg.xlsx'):
        # 如果不存在，则创建
        # 在目标文件夹下创建一个excel文件，用来存放avg_df的数据
        with pd.ExcelWriter(target_path + target_folder_name + '/avg.xlsx', engine='openpyxl') as writer:
            # 创建一个dataframe，用来存放表头数据：acc, SE, SP, PC, DC, IOU
            df = pd.DataFrame(columns=['exp_name', 'acc', 'SE', 'SP', 'PC', 'DC', 'IOU'])
            df.to_excel(writer, sheet_name='AVG', header=['exp_name', 'acc', 'SE', 'SP', 'PC', 'DC', 'IOU'], index=False)
            print('目标文件夹下不存在avg.xlsx文件，已创建')
    else:
        print('目标文件夹下已存在avg.xlsx文件')

    folders = [f for f in os.listdir(path) if os.path.isdir(os.path.join(path, f))]  # 这里是在获取文件夹列表

    # 获取文件夹创建时间并进行排序
    sorted_folders = sorted(folders, key=lambda x: os.path.getmtime(os.path.join(path, x)))



    # 用一个函数来获取源文件夹下的所有文件名
    file_name = get_folder_name(path)

    # 获取path下一级的文件夹个数
    folder_num = get_folder_num(path)



    # 依次检测file_name中每个文件是否为文件夹
    for file in sorted_folders:
        print('当前文件夹名字：', file)
        file = path + file
        # 如果是文件夹，则继续其他操作
        if os.path.isdir(file):
            # 用一个函数读取当前文件夹下的record.xlsx文件
            record = read_record(file)
            if record.empty:
                print('当前文件夹下不存在record.xlsx文件，跳过')
                continue
            else:
                # 用一个函数来检测record有没有超过100行，如果不超过100行，则跳过
                if not check_record(record):
                    print('当前文件夹下record.xlsx文件行数不足100行，跳过')
                    continue

                # 用一个函数检测record.xlsx第一列中是否已经存在名为exp_name的元素，如果存在，则跳过，函数参数为record的路径名，exp_name
                if check_exp_name(target_path + target_folder_name, file):
                    print('当前文件夹下record.xlsx文件中已存在名为%s的元素，跳过' % file)
                    continue

            # 用一个函数来读取record的第3列往后的所有列（包括第3列）的最后10行数据
            last_n_row = get_last_n_row(record, 3, 10)  # last_n_row是一个dataframe，dataframe是一个二维数组，第一维是行，第二维是列。
            # 用一个函数来计算last_n_row每一列的平均值
            avg = get_avg(last_n_row)  # avg是一个series，series是一个一维数组，第一维是列。
            # 将exp_name添加到avg的第一列
            avg = pd.concat([pd.Series([file]), avg])
            # 将avg转置
            avg = avg.T


            # 用openxyl来实现excel的写入
            # 打开目标文件夹下的avg.xlsx文件
            wb = openpyxl.load_workbook(target_path + target_folder_name + '/avg.xlsx')
            # 获取AVG的sheet
            sheet = wb['AVG']
            # 将avg转换为list
            avg_list = avg.values.tolist()
            # 将avg_list的数据写入到AVG的sheet中
            for row in avg_list:
                sheet.append(row)
            # 保存
            wb.save(target_path + target_folder_name + '/avg.xlsx')

        # 打印进度条，以及当前文件夹名字
        print('进度：', file_name.index(file) / folder_num * 100, '%')

    print('程序执行完毕')









